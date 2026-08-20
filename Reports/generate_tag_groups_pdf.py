# pyright: reportMissingTypeStubs=false

from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Attessa - Tags - Organized.xlsx"
OUTPUT = ROOT / "Reports" / "Attessa Tag Groups.pdf"
MASTER_SHEETS = {"Tags", "Alarms"}


def populated_rows(sheet):
    return [
        tuple(value for value in row)
        for row in sheet.iter_rows(values_only=True)
        if any(value not in (None, "") for value in row)
    ]


def extract_group(sheet):
    rows = populated_rows(sheet)
    if not rows:
        return []

    first = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
    has_header = "name" in first
    if has_header:
        name_index = first.index("name")
        unit_index = first.index("unit") if "unit" in first else None
        rows = rows[1:]
    else:
        first_value = str(rows[0][0] or "")
        name_index = 1 if "." in first_value else 2
        unit_index = 3 if name_index == 1 else 4

    entries = []
    for row in rows:
        if len(row) <= name_index or row[name_index] in (None, ""):
            continue
        unit = row[unit_index] if unit_index is not None and len(row) > unit_index else ""
        entries.append((str(row[name_index]).strip(), str(unit or "").strip()))
    return entries


def page_footer(canvas, document):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#B7C7CE"))
    canvas.line(0.65 * inch, 0.48 * inch, 7.85 * inch, 0.48 * inch)
    canvas.setFillColor(colors.HexColor("#49636E"))
    canvas.setFont("Helvetica", 8)
    canvas.drawString(0.65 * inch, 0.3 * inch, "Attessa Tag Groups")
    canvas.drawRightString(7.85 * inch, 0.3 * inch, f"Page {document.page}")
    canvas.restoreState()


def build_pdf():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    groups = [
        (sheet.title, extract_group(sheet))
        for sheet in workbook.worksheets
        if sheet.title not in MASTER_SHEETS
    ]
    master_counts = {
        sheet.title: len(populated_rows(sheet)) - 1
        for sheet in workbook.worksheets
        if sheet.title in MASTER_SHEETS
    }

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=24,
        leading=29,
        textColor=colors.HexColor("#14394A"),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["BodyText"],
        fontSize=10,
        leading=15,
        textColor=colors.HexColor("#49636E"),
        alignment=TA_CENTER,
        spaceAfter=18,
    )
    group_style = ParagraphStyle(
        "GroupTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=21,
        textColor=colors.HexColor("#14394A"),
        spaceAfter=4,
    )
    count_style = ParagraphStyle(
        "GroupCount",
        parent=styles["BodyText"],
        fontSize=8,
        textColor=colors.HexColor("#5D737C"),
        spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#102B36"),
    )
    unit_style = ParagraphStyle(
        "Unit",
        parent=cell_style,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#49636E"),
    )

    document = SimpleDocTemplate(
        str(OUTPUT),
        pagesize=letter,
        rightMargin=0.65 * inch,
        leftMargin=0.65 * inch,
        topMargin=0.62 * inch,
        bottomMargin=0.62 * inch,
        title="Attessa Tag Groups",
        author="Attessa Dashboard Project",
        subject="Organized workbook tab groups and tag names",
    )

    story = [
        Spacer(1, 0.75 * inch),
        Paragraph("Attessa Tag Groups", title_style),
        Paragraph(
            "Organized equipment groups and tag names from "
            "<b>Attessa - Tags - Organized.xlsx</b>",
            subtitle_style,
        ),
        Spacer(1, 0.15 * inch),
    ]

    index_rows = [["Workbook tab / group", "Names"]]
    index_rows.extend([[escape(name), str(len(entries))] for name, entries in groups])
    index_table = Table(index_rows, colWidths=[5.7 * inch, 0.8 * inch], repeatRows=1)
    index_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14394A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B7C7CE")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F6F8")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend(
        [
            index_table,
            Spacer(1, 0.18 * inch),
            Paragraph(
                f"Master sheets summarized: Tags ({master_counts.get('Tags', 0):,} rows), "
                f"Alarms ({master_counts.get('Alarms', 0):,} rows). These source sheets are not "
                "repeated because this report focuses on the organized group tabs.",
                count_style,
            ),
        ]
    )

    for group_index, (group_name, entries) in enumerate(groups):
        story.append(PageBreak())
        story.append(Paragraph(escape(group_name), group_style))
        story.append(Paragraph(f"{len(entries)} names in this group", count_style))

        table_rows = [["Name", "Unit"]]
        table_rows.extend(
            [Paragraph(escape(name), cell_style), Paragraph(escape(unit), unit_style)]
            for name, unit in entries
        )
        table = LongTable(table_rows, colWidths=[6.55 * inch, 0.65 * inch], repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#14394A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("ALIGN", (1, 0), (1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#C8D5DA")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8F9")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    document.build(story, onFirstPage=page_footer, onLaterPages=page_footer)
    print(f"Created: {OUTPUT}")
    print(f"Groups: {len(groups)}")
    print(f"Names: {sum(len(entries) for _, entries in groups)}")


if __name__ == "__main__":
    build_pdf()